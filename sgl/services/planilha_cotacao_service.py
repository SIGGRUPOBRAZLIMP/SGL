"""
SGL - Gerador Automático de Planilha de Cotação — V1.0

Layout revisado com feedback dos vendedores:
  - STATUS (VIÁVEL/INVIÁVEL) como coluna A
  - QTD única (removeu QTD MÍNIMA)
  - CÓD PROPOSTA removido
  - MÍN VENDA e MÍN TOTAL VENDA antes do Menor Custo
  - MARGEM % após OBS Vencedor
  - 15 fornecedores + Estoque + Disputa
  - Nomes dos fornecedores editáveis (células desbloqueadas)
"""
import io
import logging
import os
from datetime import datetime, timezone
from threading import Thread

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# ============================================================
# ESTILOS GLOBAIS
# ============================================================
_title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
_label_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
_value_font = Font(name='Arial', size=10)
_hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
_dat_font = Font(name='Arial', size=9)
_bld_font = Font(name='Arial', size=9, bold=True)
_ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
_thin = Border(
    left=Side('thin', color='D9D9D9'), right=Side('thin', color='D9D9D9'),
    top=Side('thin', color='D9D9D9'), bottom=Side('thin', color='D9D9D9'),
)
_LOCKED = Protection(locked=True)
_UNLOCKED = Protection(locked=False)

_fill_edital = PatternFill('solid', fgColor='1F4E79')
_fill_proposta = PatternFill('solid', fgColor='2E75B6')
_fill_calc = PatternFill('solid', fgColor='548235')
_fill_result = PatternFill('solid', fgColor='BF8F00')
_fill_formula = PatternFill('solid', fgColor='E2EFDA')
_fill_marca_v = PatternFill('solid', fgColor='D5E8D4')
_fill_resumo = PatternFill('solid', fgColor='F2F2F2')
_fill_prop_auto = PatternFill('solid', fgColor='D6E4F0')
_fill_disputa = PatternFill('solid', fgColor='8B0000')

# 15 fornecedores genéricos (nomes editáveis na planilha)
NUM_FORNECEDORES = 15
FORNECEDORES_PADRAO = [f'FORNECEDOR {i+1}' for i in range(NUM_FORNECEDORES)]
FORN_COLORS = ['4472C4', '548235', 'BF8F00', 'C00000', '7030A0']
CPF = 4  # colunas por fornecedor (Preço, Marca, Cód, Obs)
SENHA_PROTECAO = 'sgl2026'

# Posições calculadas
CS = 20                           # fornecedores começam na coluna 20 (T)
CE = CS + NUM_FORNECEDORES * CPF  # estoque começa após fornecedores
CD = CE + 4                       # disputa começa após estoque


def _fmt_date(dt):
    if not dt:
        return ''
    if isinstance(dt, str):
        return dt[:10]
    return dt.strftime('%d/%m/%Y')


def _fmt_currency(val):
    if not val:
        return ''
    try:
        return f'R$ {float(val):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return str(val)


# ============================================================
# ABA: DADOS DO EDITAL
# ============================================================
def _criar_aba_dados_edital(wb, edital_dict):
    ws = wb.active
    ws.title = "Dados do Edital"
    ws.merge_cells('A1:D1')
    ws['A1'] = 'DADOS DO EDITAL'
    ws['A1'].font = _title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 60

    campos = [
        ('Órgão', edital_dict.get('orgao_razao_social', '')),
        ('CNPJ', edital_dict.get('orgao_cnpj', '')),
        ('UF / Município', f"{edital_dict.get('uf', '')} / {edital_dict.get('municipio', '')}"),
        ('Modalidade', edital_dict.get('modalidade_nome', '')),
        ('Nº Processo', edital_dict.get('numero_processo', '')),
        ('Nº Pregão', edital_dict.get('numero_pregao', '')),
        ('Plataforma', (edital_dict.get('plataforma_origem', '') or '').upper()),
        ('SRP', 'Sim' if edital_dict.get('srp') else 'Não'),
        ('Data Publicação', _fmt_date(edital_dict.get('data_publicacao'))),
        ('Data Abertura Propostas', _fmt_date(edital_dict.get('data_abertura_proposta'))),
        ('Data Encerramento', _fmt_date(edital_dict.get('data_encerramento_proposta'))),
        ('Data Certame', _fmt_date(edital_dict.get('data_certame'))),
        ('Valor Estimado Total', _fmt_currency(edital_dict.get('valor_estimado'))),
        ('Status', edital_dict.get('status', '')),
        ('Link Edital', edital_dict.get('url_original', '')),
        ('Objeto', edital_dict.get('objeto_resumo') or edital_dict.get('objeto_completo', '')),
    ]

    for i, (lbl, val) in enumerate(campos, 3):
        ws.cell(row=i, column=1, value=lbl).font = _label_font
        ws.cell(row=i, column=1).border = _thin
        ws.cell(row=i, column=1).protection = _UNLOCKED
        c = ws.cell(row=i, column=2, value=val)
        c.font = _value_font
        c.border = _thin
        c.protection = _UNLOCKED
        if lbl == 'Objeto':
            c.alignment = Alignment(wrap_text=True, vertical='top')

    # --- RESUMO COTAÇÃO ---
    r = len(campos) + 5
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = 'RESUMO DA COTAÇÃO (automático)'
    ws[f'A{r}'].font = _title_font
    ws[f'A{r}'].alignment = Alignment(horizontal='center')
    ws[f'A{r}'].protection = _LOCKED

    for i, (lbl, f) in enumerate([
        ('Total de Itens', "=COUNTA('Cotação'!C2:C500)"),
        ('Valor Total Máximo (edital)', "=SUM('Cotação'!I2:I500)"),
        ('Valor Total Proposta (c/ margem)', "=SUM('Cotação'!M2:M500)"),
        ('Itens Viáveis', "=COUNTIF('Cotação'!A2:A500,\"VIÁVEL\")"),
        ('Itens Inviáveis', "=COUNTIF('Cotação'!A2:A500,\"INVIÁVEL\")"),
    ], r + 2):
        cl = ws.cell(row=i, column=1, value=lbl)
        cl.font = _label_font
        cl.border = _thin
        cl.fill = _fill_resumo
        cl.protection = _LOCKED
        cv = ws.cell(row=i, column=2, value=f)
        cv.font = Font(name='Arial', bold=True, size=11)
        cv.number_format = '#,##0.00'
        cv.border = _thin
        cv.fill = _fill_formula
        cv.protection = _LOCKED

    # --- RESUMO DISPUTA ---
    sd_col = get_column_letter(CD)
    tf_col = get_column_letter(CD + 2)

    rd = r + 9
    ws.merge_cells(f'A{rd}:D{rd}')
    ws[f'A{rd}'] = 'RESUMO DA DISPUTA (automático)'
    ws[f'A{rd}'].font = _title_font
    ws[f'A{rd}'].alignment = Alignment(horizontal='center')
    ws[f'A{rd}'].protection = _LOCKED

    for i, (lbl, f) in enumerate([
        ('Itens Vencidos',
         f"=COUNTIF('Cotação'!{sd_col}2:{sd_col}500,\"VENCIDO\")"),
        ('Itens Não Vencidos',
         f"=COUNTIF('Cotação'!{sd_col}2:{sd_col}500,\"NÃO VENCIDO\")"),
        ('Itens Desertos/Fracassados',
         f"=COUNTIF('Cotação'!{sd_col}2:{sd_col}500,\"DESERTO\")+COUNTIF('Cotação'!{sd_col}2:{sd_col}500,\"FRACASSADO\")"),
        ('Valor Total Final (Vencidos)',
         f"=SUMPRODUCT(('Cotação'!{sd_col}2:{sd_col}500=\"VENCIDO\")*('Cotação'!{tf_col}2:{tf_col}500))"),
    ], rd + 2):
        cl = ws.cell(row=i, column=1, value=lbl)
        cl.font = _label_font
        cl.border = _thin
        cl.fill = _fill_resumo
        cl.protection = _LOCKED
        cv = ws.cell(row=i, column=2, value=f)
        cv.font = Font(name='Arial', bold=True, size=11)
        cv.number_format = '#,##0.00'
        cv.border = _thin
        cv.fill = _fill_formula
        cv.protection = _LOCKED

    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    return ws


# ============================================================
# ABA: COTAÇÃO (principal) — V1.0
# ============================================================
def _criar_aba_cotacao(wb, itens, fornecedores=None):
    ws = wb.create_sheet("Cotação")
    forns = fornecedores or FORNECEDORES_PADRAO

    # === CABEÇALHOS FIXOS (A..S) ===
    hdrs = [
        (1,  'STATUS',              10, _fill_result,   True),
        (2,  'LOTE/GRUPO',           8, _fill_edital,   False),
        (3,  'ITEM',                 6, _fill_edital,   False),
        (4,  'ESPECIFICAÇÃO',       55, _fill_edital,   False),
        (5,  'CATMAT',              10, _fill_edital,   False),
        (6,  'QTD',                 10, _fill_edital,   False),
        (7,  'UNID',                 7, _fill_edital,   False),
        (8,  'PREÇO UNIT\nMÁXIMO', 12, _fill_edital,   False),
        (9,  'PREÇO TOTAL\nMÁXIMO', 13, _fill_edital,  True),
        (10, 'MARCA\nPROPOSTA',    14, _fill_proposta,  True),
        (11, 'MODELO',              12, _fill_proposta,  False),
        (12, 'MÍN\nVENDA',         11, _fill_result,    True),
        (13, 'MÍN TOTAL\nVENDA',   12, _fill_result,    True),
        (14, 'MENOR\nCUSTO',       11, _fill_calc,      True),
        (15, 'FORN.\nVENCEDOR',    16, _fill_calc,      True),
        (16, 'MARCA\nVENCEDORA',   14, PatternFill('solid', fgColor='375623'), True),
        (17, 'CÓD\nVENCEDOR',     10, PatternFill('solid', fgColor='375623'), True),
        (18, 'OBS\nVENCEDOR',     16, PatternFill('solid', fgColor='375623'), True),
        (19, 'MARGEM\n%',           9, _fill_result,    True),
    ]

    formula_cols = set()
    for ci, txt, w, fill, is_f in hdrs:
        c = ws.cell(row=1, column=ci, value=txt)
        c.font = _hdr_font
        c.alignment = _ac
        c.fill = fill
        c.border = _thin
        c.protection = _LOCKED
        ws.column_dimensions[get_column_letter(ci)].width = w
        if is_f:
            formula_cols.add(ci)

    # === CABEÇALHOS FORNECEDORES (15 × 4) ===
    cs = CS
    for fi, fn in enumerate(forns):
        color = FORN_COLORS[fi % len(FORN_COLORS)]
        for di, (lbl, w) in enumerate([
            (f'{fn}\nPREÇO', 11), ('MARCA', 10), ('CÓD', 8), ('OBS', 14)
        ]):
            col = cs + fi * CPF + di
            c = ws.cell(row=1, column=col, value=lbl)
            c.font = _hdr_font
            c.alignment = _ac
            c.border = _thin
            c.fill = PatternFill('solid', fgColor=color)
            c.protection = _UNLOCKED  # editável para renomear fornecedor
            ws.column_dimensions[get_column_letter(col)].width = w

    # === ESTOQUE ===
    ce = CE
    for di, (lbl, w) in enumerate([
        ('ESTOQUE\nPREÇO', 11), ('ESTOQUE\nMARCA', 10),
        ('ESTOQUE\nCÓDIGO', 10), ('ESTOQUE\nOBS', 14)
    ]):
        col = ce + di
        c = ws.cell(row=1, column=col, value=lbl)
        c.font = _hdr_font
        c.alignment = _ac
        c.border = _thin
        c.fill = PatternFill('solid', fgColor='C65911')
        c.protection = _LOCKED
        ws.column_dimensions[get_column_letter(col)].width = w

    # === DISPUTA ===
    cd = CD
    disputa_hdrs = [
        ('STATUS\nDISPUTA', 13, False),
        ('PREÇO\nFINAL', 12, False),
        ('TOTAL\nFINAL', 12, True),
        ('DESCONTO\n%', 9, True),
        ('OBS\nDISPUTA', 18, False),
    ]
    disputa_formula_cols = set()
    for di, (lbl, w, is_f) in enumerate(disputa_hdrs):
        col = cd + di
        c = ws.cell(row=1, column=col, value=lbl)
        c.font = _hdr_font
        c.alignment = _ac
        c.border = _thin
        c.fill = _fill_disputa
        c.protection = _LOCKED
        ws.column_dimensions[get_column_letter(col)].width = w
        if is_f:
            disputa_formula_cols.add(col)

    # Letras disputa
    pf_letter = get_column_letter(cd + 1)
    tf_letter = get_column_letter(cd + 2)

    # === LETRAS REFERÊNCIA FORNECEDORES ===
    fpl, fml, fcl, fol = [], [], [], []
    for fi in range(len(forns)):
        base = cs + fi * CPF
        fpl.append(get_column_letter(base))
        fml.append(get_column_letter(base + 1))
        fcl.append(get_column_letter(base + 2))
        fol.append(get_column_letter(base + 3))

    epl = get_column_letter(ce)
    eml = get_column_letter(ce + 1)
    ecl = get_column_letter(ce + 2)
    eol = get_column_letter(ce + 3)

    all_price = fpl + [epl]
    all_marca = fml + [eml]
    all_cod = fcl + [ecl]
    all_obs = fol + [eol]

    def _build_nested(parts):
        nested = '""'
        for p in reversed(parts):
            nested = f'{p},{nested})'
        return nested

    # === PREENCHER ITENS + FÓRMULAS ===
    total_rows = max(len(itens) + 5, 50)
    total_rows = min(total_rows, 500)

    for row in range(2, total_rows + 2):
        r = str(row)
        item_idx = row - 2

        if item_idx < len(itens):
            item = itens[item_idx]
            ws.cell(row=row, column=3, value=item.get('numero_item', item_idx + 1))
            ws.cell(row=row, column=4, value=item.get('descricao', ''))
            ws.cell(row=row, column=5, value=item.get('codigo_referencia', ''))
            ws.cell(row=row, column=6, value=item.get('quantidade') or '')
            ws.cell(row=row, column=7, value=item.get('unidade_compra', ''))
            preco = item.get('preco_unitario_maximo')
            if preco:
                try:
                    ws.cell(row=row, column=8, value=float(preco))
                except (ValueError, TypeError):
                    ws.cell(row=row, column=8, value=preco)
            ws.cell(row=row, column=2, value=item.get('grupo_lote', ''))

        # A = STATUS
        ws.cell(row=row, column=1,
                value=f'=IF(L{r}="","",IF(L{r}<=H{r},"VIÁVEL","INVIÁVEL"))')

        # I = PREÇO TOTAL MÁXIMO
        ws.cell(row=row, column=9,
                value=f'=IF(H{r}<>"",H{r}*F{r},"")').number_format = '#,##0.00'

        # N = MENOR CUSTO
        count_p = '+'.join([f'({c}{r}>0)*1' for c in all_price])
        min_p = ','.join([f'IF({c}{r}>0,{c}{r},9999999)' for c in all_price])
        ws.cell(row=row, column=14,
                value=f'=IF(AND(C{r}<>"",({count_p})>0),MIN({min_p}),"")').number_format = '#,##0.00'
        ws.cell(row=row, column=14).font = _bld_font

        # L = MÍN VENDA
        ws.cell(row=row, column=12,
                value=f"=IF(N{r}<>\"\",ROUND(N{r}*(1+'Configuração'!$B$3/100),2),\"\")").number_format = '#,##0.00'

        # M = MÍN TOTAL VENDA
        ws.cell(row=row, column=13,
                value=f'=IF(L{r}<>"",L{r}*F{r},"")').number_format = '#,##0.00'

        # O = FORN. VENCEDOR
        pn = [
            f'IF({all_price[fi]}{r}=N{r},LEFT({all_price[fi]}$1,FIND(CHAR(10),{all_price[fi]}$1&CHAR(10))-1)'
            for fi in range(len(all_price))
        ]
        ws.cell(row=row, column=15,
                value=f'=IF(N{r}<>"",{_build_nested(pn)},"")').font = _bld_font

        # P = MARCA VENCEDORA
        pm = [f'IF({all_price[fi]}{r}=N{r},{all_marca[fi]}{r}'
              for fi in range(len(all_price))]
        ws.cell(row=row, column=16,
                value=f'=IF(N{r}<>"",{_build_nested(pm)},"")').font = _bld_font

        # Q = CÓD VENCEDOR
        pc = [f'IF({all_price[fi]}{r}=N{r},{all_cod[fi]}{r}'
              for fi in range(len(all_price))]
        ws.cell(row=row, column=17,
                value=f'=IF(N{r}<>"",{_build_nested(pc)},"")').font = _bld_font

        # R = OBS VENCEDOR
        po = [f'IF({all_price[fi]}{r}=N{r},{all_obs[fi]}{r}'
              for fi in range(len(all_price))]
        ws.cell(row=row, column=18,
                value=f'=IF(N{r}<>"",{_build_nested(po)},"")').font = _bld_font

        # J = MARCA PROPOSTA
        ws.cell(row=row, column=10,
                value=f'=IF(P{r}<>"",P{r},"")').font = _bld_font

        # S = MARGEM %
        ws.cell(row=row, column=19,
                value=f'=IF(AND(L{r}<>"",H{r}<>"",L{r}>0),ROUND((H{r}/L{r}-1)*100,1),"")').number_format = '0.0'

        # DISPUTA: TOTAL FINAL
        ws.cell(row=row, column=cd + 2,
                value=f'=IF({pf_letter}{r}<>"",{pf_letter}{r}*F{r},"")').number_format = '#,##0.00'
        # DISPUTA: DESCONTO %
        ws.cell(row=row, column=cd + 3,
                value=f'=IF(AND({pf_letter}{r}<>"",L{r}<>"",L{r}>0),ROUND((1-{pf_letter}{r}/L{r})*100,1),"")').number_format = '0.0'

        # === FORMATAÇÃO ===
        last_col_fmt = cd + len(disputa_hdrs)
        for col in range(1, last_col_fmt + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin
            cell.font = _dat_font
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 4))
            if col in (8, 9, 12, 13, 14):
                cell.number_format = '#,##0.00'
            if cs <= col < ce and (col - cs) % CPF == 0:
                cell.number_format = '#,##0.00'
            if col == ce:
                cell.number_format = '#,##0.00'
            if col in (cd + 1, cd + 2):
                cell.number_format = '#,##0.00'
            if col in formula_cols or col in disputa_formula_cols:
                cell.protection = _LOCKED
                cell.fill = _fill_formula
            else:
                cell.protection = _UNLOCKED

        for c in (16, 17, 18):
            ws.cell(row=row, column=c).fill = _fill_marca_v
        ws.cell(row=row, column=10).fill = _fill_prop_auto

    # === DROPDOWN STATUS DISPUTA ===
    dv = DataValidation(
        type="list",
        formula1='"VENCIDO,NÃO VENCIDO,DESERTO,FRACASSADO"',
        allow_blank=True,
    )
    dv.error = "Selecione: VENCIDO, NÃO VENCIDO, DESERTO ou FRACASSADO"
    dv.errorTitle = "Status Inválido"
    ws.add_data_validation(dv)
    dv.add(f'{get_column_letter(cd)}2:{get_column_letter(cd)}{total_rows + 1}')

    # === TOTAIS ===
    rt = total_rows + 2
    ws.cell(row=rt, column=3, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=rt, column=3).protection = _LOCKED
    for ci in (9, 13):
        L = get_column_letter(ci)
        c = ws.cell(row=rt, column=ci, value=f'=SUM({L}2:{L}{rt - 1})')
        c.number_format = '#,##0.00'
        c.font = Font(name='Arial', bold=True, size=11)
        c.protection = _LOCKED
        c.fill = _fill_formula
    c = ws.cell(row=rt, column=cd + 2,
                value=f'=SUM({tf_letter}2:{tf_letter}{rt - 1})')
    c.number_format = '#,##0.00'
    c.font = Font(name='Arial', bold=True, size=11)
    c.protection = _LOCKED
    c.fill = _fill_formula

    ws.freeze_panes = 'E2'
    last_letter = get_column_letter(cd + len(disputa_hdrs) - 1)
    ws.auto_filter.ref = f'A1:{last_letter}{rt - 1}'

    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    return ws


# ============================================================
# ABA: CONFIGURAÇÃO
# ============================================================
def _criar_aba_config(wb):
    wc = wb.create_sheet("Configuração")
    wc['A1'] = 'CONFIGURAÇÕES'
    wc['A1'].font = _title_font
    wc.column_dimensions['A'].width = 35
    wc.column_dimensions['B'].width = 50

    for i, (lbl, val) in enumerate([
        ('Margem Padrão (%)', 30),
        ('Empresa', 'DISTRIB BRAZLIMP LTDA'),
        ('CNPJ', ''), ('Responsável', ''), ('Telefone', ''), ('E-mail', ''),
    ], 3):
        wc.cell(row=i, column=1, value=lbl).font = _label_font
        wc.cell(row=i, column=2, value=val).font = Font(name='Arial', size=10, color='0000FF')

    wc.cell(row=10, column=1, value='PROTEÇÃO DA PLANILHA').font = Font(
        name='Arial', bold=True, size=13, color='1F4E79')

    wc.merge_cells('A12:A13')
    c = wc.cell(row=12, column=1, value='🔓 DESPROTEGER')
    c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='C00000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    wc.cell(row=12, column=2, value='Revisão → Desproteger Planilha → Senha: sgl2026').font = _value_font

    wc.merge_cells('A15:A16')
    c = wc.cell(row=15, column=1, value='🔒 PROTEGER')
    c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='548235')
    c.alignment = Alignment(horizontal='center', vertical='center')
    wc.cell(row=15, column=2, value='Revisão → Proteger Planilha → Senha: sgl2026').font = _value_font

    wc.cell(row=18, column=1, value='Senha: sgl2026').font = Font(
        name='Arial', bold=True, size=12, color='C00000')
    wc.cell(row=18, column=1).fill = PatternFill('solid', fgColor='FFF2CC')

    instrucoes = [
        '1. Itens preenchidos automaticamente pela IA',
        '2. Preencha preço, marca, código e OBS de cada fornecedor',
        '3. MARCA PROPOSTA puxa do fornecedor vencedor automaticamente',
        '4. Para alterar: Revisão → Desproteger → senha sgl2026',
        '5. Colunas VERDE CLARO = fórmulas protegidas',
        '6. Colunas AZUL CLARO = proposta (auto do vencedor)',
        '7. Margem padrão — alterável na célula B3 acima',
        '8. VIÁVEL = mín venda ≤ preço máximo do edital',
        '9. Nomes dos fornecedores são editáveis no cabeçalho',
        '10. Para ocultar colunas vazias: selecionar → botão direito → Ocultar',
        '',
        'RESULTADO DA DISPUTA:',
        '11. Após a sessão, preencha STATUS DISPUTA (dropdown)',
        '12. Preencha PREÇO FINAL com valor efetivo do lance',
        '13. TOTAL FINAL e DESCONTO % são calculados automaticamente',
        '14. Itens VENCIDOS serão usados para gerar planilha reajustada',
    ]
    wc.cell(row=20, column=1, value='INSTRUÇÕES').font = _title_font
    for i, t in enumerate(instrucoes, 22):
        wc.cell(row=i, column=1, value=t).font = Font(name='Arial', size=9, color='333333')
    return wc


# ============================================================
# FUNÇÃO PRINCIPAL: GERAR PLANILHA
# ============================================================
def gerar_planilha_cotacao(edital_id, app=None):
    if app is None:
        from flask import current_app
        app = current_app._get_current_object()

    with app.app_context():
        from ..models.database import db, Edital, ItemEditalExtraido, EditalArquivo

        edital = Edital.query.get(edital_id)
        if not edital:
            logger.error("Edital %d não encontrado para gerar planilha", edital_id)
            return None

        edital_dict = edital.to_dict()
        itens_db = ItemEditalExtraido.query.filter_by(edital_id=edital_id).order_by(
            ItemEditalExtraido.numero_item.asc()
        ).all()

        itens = []
        for item in itens_db:
            itens.append({
                'numero_item': item.numero_item,
                'descricao': item.descricao,
                'codigo_referencia': item.codigo_referencia,
                'quantidade': float(item.quantidade) if item.quantidade else None,
                'unidade_compra': item.unidade_compra or 'UN',
                'preco_unitario_maximo': float(item.preco_unitario_maximo) if item.preco_unitario_maximo else None,
                'grupo_lote': item.grupo_lote,
            })

        logger.info("Gerando planilha cotação V1.0: edital=%d, %d itens", edital_id, len(itens))

        wb = Workbook()
        _criar_aba_dados_edital(wb, edital_dict)
        _criar_aba_cotacao(wb, itens)
        _criar_aba_config(wb)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_bytes = buffer.getvalue()

        logger.info("Planilha cotação V1.0 gerada: edital=%d, %d bytes, %d itens",
                     edital_id, len(xlsx_bytes), len(itens))
        return xlsx_bytes


def gerar_e_enviar_planilha(edital_id, app=None):
    if app is None:
        from flask import current_app
        app = current_app._get_current_object()

    with app.app_context():
        from ..models.database import db, Edital, EditalArquivo
        from . import dropbox_service

        edital = Edital.query.get(edital_id)
        if not edital:
            logger.error("Edital %d não encontrado", edital_id)
            return

        xlsx_bytes = gerar_planilha_cotacao(edital_id, app)
        if not xlsx_bytes:
            logger.warning("Falha ao gerar planilha para edital %d", edital_id)
            return

        orgao_curto = (edital.orgao_razao_social or 'Edital')[:40].replace('/', '-')
        nome_arquivo = f"COTACAO_{edital.id}_{orgao_curto}.xlsx"

        try:
            pasta = dropbox_service.gerar_pasta_edital(edital)
            dropbox_service.criar_pasta(pasta)
            resultado = dropbox_service.upload_arquivo(
                xlsx_bytes, f"{pasta}/{nome_arquivo}", nome_arquivo=nome_arquivo)

            if resultado:
                existente = EditalArquivo.query.filter_by(
                    edital_id=edital.id, tipo='cotacao').first()
                if existente:
                    existente.url_cloudinary = resultado.get('shared_link') or resultado['dropbox_path']
                    existente.nome_arquivo = nome_arquivo
                    existente.tamanho_bytes = resultado['tamanho']
                else:
                    arquivo = EditalArquivo(
                        edital_id=edital.id, tipo='cotacao', nome_arquivo=nome_arquivo,
                        url_cloudinary=resultado.get('shared_link') or resultado['dropbox_path'],
                        tamanho_bytes=resultado['tamanho'],
                        mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    db.session.add(arquivo)
                db.session.commit()
                logger.info("Planilha cotação V1.0 enviada: edital=%d, %s", edital_id, nome_arquivo)
            else:
                logger.warning("Falha upload Dropbox planilha edital %d", edital_id)
        except Exception as e:
            db.session.rollback()
            logger.error("Erro enviar planilha Dropbox edital %d: %s", edital_id, e)


def disparar_geracao_planilha_async(edital_id, app):
    thread = Thread(target=gerar_e_enviar_planilha, args=(edital_id, app), daemon=True)
    thread.start()
    logger.info("Geração planilha V1.0 assíncrona disparada para edital %d", edital_id)
    return thread
