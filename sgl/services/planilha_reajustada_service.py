"""
SGL - Gerador de Planilha Reajustada (Pós-Disputa)

Gera XLSX com itens VENCIDOS na disputa + controle de entregas e reajustes.
Fluxo: resultado da disputa registrado → gera planilha reajustada → Dropbox.

Abas:
  1. Contrato — dados do contrato, órgão, vigência
  2. Itens Vencidos — itens com status VENCIDO + preço final
  3. Entregas — controle de entregas parciais (vazio, preenchido manualmente)
  4. Reajustes — histórico de reajustes de preço (vazio, preenchido manualmente)
"""
import io
import logging
from datetime import datetime, timezone
from threading import Thread

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# ============================================================
# ESTILOS
# ============================================================
_title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
_subtitle_font = Font(name='Arial', bold=True, size=12, color='1F4E79')
_label_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
_value_font = Font(name='Arial', size=10)
_hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
_dat_font = Font(name='Arial', size=9)
_bld_font = Font(name='Arial', size=9, bold=True)
_ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
_thin = Border(
    left=Side('thin', color='D9D9D9'), right=Side('thin', color='D9D9D9'),
    top=Side('thin', color='D9D9D9'), bottom=Side('thin', color='D9D9D9'),
)
_LOCKED = Protection(locked=True)
_UNLOCKED = Protection(locked=False)

_fill_hdr = PatternFill('solid', fgColor='1F4E79')
_fill_ok = PatternFill('solid', fgColor='548235')
_fill_formula = PatternFill('solid', fgColor='E2EFDA')
_fill_resumo = PatternFill('solid', fgColor='F2F2F2')
_fill_entrega = PatternFill('solid', fgColor='2E75B6')
_fill_reajuste = PatternFill('solid', fgColor='BF8F00')

SENHA_PROTECAO = 'sgl2026'


def _fmt_date(dt):
    if not dt:
        return ''
    if isinstance(dt, str):
        return dt[:10]
    return dt.strftime('%d/%m/%Y')


def _fmt_currency(val):
    if not val:
        return ''
    try:
        return f'R$ {float(val):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return str(val)


# ============================================================
# ABA 1: CONTRATO
# ============================================================
def _criar_aba_contrato(wb, edital_dict, itens_vencidos):
    """Dados do contrato + resumo dos itens vencidos."""
    ws = wb.active
    ws.title = "Contrato"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60

    ws.merge_cells('A1:B1')
    ws['A1'] = 'PLANILHA REAJUSTADA — CONTRATO'
    ws['A1'].font = _title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    campos = [
        ('Nº Contrato', ''),  # preenchido manualmente
        ('Órgão', edital_dict.get('orgao_razao_social', '')),
        ('CNPJ Órgão', edital_dict.get('orgao_cnpj', '')),
        ('UF / Município', f"{edital_dict.get('uf', '')} / {edital_dict.get('municipio', '')}"),
        ('Modalidade', edital_dict.get('modalidade_nome', '')),
        ('Nº Processo', edital_dict.get('numero_processo', '')),
        ('Nº Pregão', edital_dict.get('numero_pregao', '')),
        ('Plataforma', (edital_dict.get('plataforma_origem', '') or '').upper()),
        ('SRP', 'Sim' if edital_dict.get('srp') else 'Não'),
        ('Data Publicação', _fmt_date(edital_dict.get('data_publicacao'))),
        ('Data Certame', _fmt_date(edital_dict.get('data_certame'))),
        ('Vigência Início', ''),   # preenchido manualmente
        ('Vigência Fim', ''),      # preenchido manualmente
        ('Valor Estimado (Edital)', _fmt_currency(edital_dict.get('valor_estimado'))),
        ('Link Edital', edital_dict.get('url_original', '')),
        ('Objeto', edital_dict.get('objeto_resumo') or edital_dict.get('objeto_completo', '')),
    ]

    for i, (lbl, val) in enumerate(campos, 3):
        cl = ws.cell(row=i, column=1, value=lbl)
        cl.font = _label_font
        cl.border = _thin
        cv = ws.cell(row=i, column=2, value=val)
        cv.font = _value_font
        cv.border = _thin
        if lbl == 'Objeto':
            cv.alignment = Alignment(wrap_text=True, vertical='top')

    # --- RESUMO VENCIDOS ---
    r = len(campos) + 5
    ws.merge_cells(f'A{r}:B{r}')
    ws[f'A{r}'] = 'RESUMO DOS ITENS VENCIDOS'
    ws[f'A{r}'].font = _subtitle_font
    ws[f'A{r}'].alignment = Alignment(horizontal='center')

    total_valor = sum(
        float(it.get('preco_total_final') or 0) for it in itens_vencidos
    )

    resumo = [
        ('Total Itens Vencidos', len(itens_vencidos)),
        ('Valor Total Adjudicado', _fmt_currency(total_valor)),
    ]

    for i, (lbl, val) in enumerate(resumo, r + 2):
        cl = ws.cell(row=i, column=1, value=lbl)
        cl.font = _label_font
        cl.border = _thin
        cl.fill = _fill_resumo
        cv = ws.cell(row=i, column=2, value=val)
        cv.font = Font(name='Arial', bold=True, size=11)
        cv.border = _thin
        cv.fill = _fill_formula

    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws


# ============================================================
# ABA 2: ITENS VENCIDOS
# ============================================================
def _criar_aba_itens_vencidos(wb, itens_vencidos):
    """Tabela com todos os itens VENCIDOS e preço final."""
    ws = wb.create_sheet("Itens Vencidos")

    hdrs = [
        ('ITEM', 6),
        ('ESPECIFICAÇÃO', 55),
        ('CATMAT', 10),
        ('QTD', 10),
        ('UNID', 7),
        ('PREÇO UNIT\nMÁXIMO (EDITAL)', 14),
        ('PREÇO UNIT\nFINAL (DISPUTA)', 14),
        ('TOTAL FINAL', 14),
        ('DESCONTO %', 10),
        ('MARCA', 14),
        ('CÓD PRODUTO', 12),
        ('OBS', 20),
    ]

    for ci, (txt, w) in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=txt)
        c.font = _hdr_font
        c.alignment = _ac
        c.fill = _fill_ok
        c.border = _thin
        c.protection = _LOCKED
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, item in enumerate(itens_vencidos, 2):
        r = str(ri)
        preco_max = item.get('preco_unitario_maximo')
        preco_final = item.get('preco_final')
        qtd = item.get('quantidade')

        ws.cell(row=ri, column=1, value=item.get('numero_item', ri - 1))
        ws.cell(row=ri, column=2, value=item.get('descricao', ''))
        ws.cell(row=ri, column=3, value=item.get('codigo_referencia', ''))

        if qtd:
            ws.cell(row=ri, column=4, value=float(qtd))
        ws.cell(row=ri, column=5, value=item.get('unidade_compra', ''))

        if preco_max:
            ws.cell(row=ri, column=6, value=float(preco_max)).number_format = '#,##0.00'
        if preco_final:
            ws.cell(row=ri, column=7, value=float(preco_final)).number_format = '#,##0.00'

        # H = total final (fórmula)
        ws.cell(row=ri, column=8, value=f'=IF(G{r}<>"",G{r}*D{r},"")').number_format = '#,##0.00'

        # I = desconto % (fórmula)
        ws.cell(
            row=ri, column=9,
            value=f'=IF(AND(G{r}<>"",F{r}<>"",F{r}>0),ROUND((1-G{r}/F{r})*100,1),"")'
        ).number_format = '0.0'

        ws.cell(row=ri, column=10, value=item.get('marca', ''))
        ws.cell(row=ri, column=11, value=item.get('codigo_produto', ''))
        ws.cell(row=ri, column=12, value=item.get('obs_disputa', ''))

        # Formatação
        for col in range(1, len(hdrs) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = _thin
            cell.font = _dat_font
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 2))

    # --- TOTAIS ---
    rt = len(itens_vencidos) + 2
    ws.cell(row=rt, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    for ci in (8,):
        L = get_column_letter(ci)
        c = ws.cell(row=rt, column=ci, value=f'=SUM({L}2:{L}{rt - 1})')
        c.number_format = '#,##0.00'
        c.font = Font(name='Arial', bold=True, size=11)
        c.fill = _fill_formula
        c.border = _thin

    ws.freeze_panes = 'C2'
    last_letter = get_column_letter(len(hdrs))
    ws.auto_filter.ref = f'A1:{last_letter}{rt - 1}'

    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws


# ============================================================
# ABA 3: ENTREGAS
# ============================================================
def _criar_aba_entregas(wb, itens_vencidos):
    """Controle de entregas parciais — preenchido manualmente pelo usuário."""
    ws = wb.create_sheet("Entregas")

    hdrs = [
        ('ITEM', 6),
        ('ESPECIFICAÇÃO', 40),
        ('QTD\nCONTRATADA', 10),
        ('ENTREGA 1\nDATA', 12),
        ('ENTREGA 1\nQTD', 10),
        ('ENTREGA 1\nNF', 12),
        ('ENTREGA 2\nDATA', 12),
        ('ENTREGA 2\nQTD', 10),
        ('ENTREGA 2\nNF', 12),
        ('ENTREGA 3\nDATA', 12),
        ('ENTREGA 3\nQTD', 10),
        ('ENTREGA 3\nNF', 12),
        ('TOTAL\nENTREGUE', 10),
        ('SALDO\nPENDENTE', 10),
        ('STATUS', 12),
    ]

    for ci, (txt, w) in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=txt)
        c.font = _hdr_font
        c.alignment = _ac
        c.fill = _fill_entrega
        c.border = _thin
        c.protection = _LOCKED
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, item in enumerate(itens_vencidos, 2):
        r = str(ri)
        ws.cell(row=ri, column=1, value=item.get('numero_item', ri - 1))
        ws.cell(row=ri, column=2, value=(item.get('descricao', '') or '')[:80])

        qtd = item.get('quantidade')
        if qtd:
            ws.cell(row=ri, column=3, value=float(qtd))

        # M = total entregue (soma das 3 entregas)
        ws.cell(row=ri, column=13, value=f'=E{r}+H{r}+K{r}').number_format = '#,##0'

        # N = saldo pendente
        ws.cell(row=ri, column=14, value=f'=IF(C{r}<>"",C{r}-M{r},"")').number_format = '#,##0'

        # O = status
        ws.cell(
            row=ri, column=15,
            value=f'=IF(C{r}="","",IF(N{r}<=0,"COMPLETA",IF(M{r}>0,"PARCIAL","PENDENTE")))'
        )

        for col in range(1, len(hdrs) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = _thin
            cell.font = _dat_font
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 2))
            if col in (13, 14, 15):
                cell.protection = _LOCKED
                cell.fill = _fill_formula
            else:
                cell.protection = _UNLOCKED

    # Validation para status
    dv = DataValidation(
        type="list",
        formula1='"PENDENTE,PARCIAL,COMPLETA"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)

    ws.freeze_panes = 'C2'
    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.autoFilter = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws


# ============================================================
# ABA 4: REAJUSTES
# ============================================================
def _criar_aba_reajustes(wb, itens_vencidos):
    """Histórico de reajustes de preço — preenchido manualmente."""
    ws = wb.create_sheet("Reajustes")

    # --- CABEÇALHO DE CONFIGURAÇÃO ---
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    ws.merge_cells('A1:B1')
    ws['A1'] = 'CONTROLE DE REAJUSTES'
    ws['A1'].font = _title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    config = [
        ('Índice de Reajuste', 'IPCA'),
        ('Periodicidade', 'Anual'),
        ('Data Base', ''),
    ]
    for i, (lbl, val) in enumerate(config, 3):
        ws.cell(row=i, column=1, value=lbl).font = _label_font
        ws.cell(row=i, column=2, value=val).font = _value_font

    # --- TABELA DE REAJUSTES ---
    r_start = 8
    ws.merge_cells(f'A{r_start}:H{r_start}')
    ws[f'A{r_start}'] = 'HISTÓRICO DE REAJUSTES'
    ws[f'A{r_start}'].font = _subtitle_font
    ws[f'A{r_start}'].alignment = Alignment(horizontal='center')

    hdrs = [
        ('Nº REAJUSTE', 10),
        ('DATA', 12),
        ('ÍNDICE\nAPLICADO (%)', 12),
        ('VALOR CONTRATO\nANTERIOR', 16),
        ('VALOR CONTRATO\nNOVO', 16),
        ('DIFERENÇA', 14),
        ('PORTARIA/\nAUTORIZAÇÃO', 18),
        ('OBS', 20),
    ]

    hr = r_start + 1
    for ci, (txt, w) in enumerate(hdrs, 1):
        c = ws.cell(row=hr, column=ci, value=txt)
        c.font = _hdr_font
        c.alignment = _ac
        c.fill = _fill_reajuste
        c.border = _thin
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 10 linhas vazias para preenchimento
    for ri in range(hr + 1, hr + 11):
        r = str(ri)
        ws.cell(row=ri, column=1, value=ri - hr)  # nº sequencial

        # F = diferença (fórmula)
        ws.cell(row=ri, column=6, value=f'=IF(AND(D{r}<>"",E{r}<>""),E{r}-D{r},"")').number_format = '#,##0.00'

        for col in range(1, len(hdrs) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = _thin
            cell.font = _dat_font
            if col in (4, 5, 6):
                cell.number_format = '#,##0.00'
            if col == 6:
                cell.protection = _LOCKED
                cell.fill = _fill_formula
            else:
                cell.protection = _UNLOCKED

    # --- TABELA DE PREÇOS POR ITEM (reajustados) ---
    ir_start = hr + 13
    ws.merge_cells(f'A{ir_start}:F{ir_start}')
    ws[f'A{ir_start}'] = 'PREÇOS UNITÁRIOS REAJUSTADOS POR ITEM'
    ws[f'A{ir_start}'].font = _subtitle_font
    ws[f'A{ir_start}'].alignment = Alignment(horizontal='center')

    item_hdrs = [
        ('ITEM', 6),
        ('ESPECIFICAÇÃO', 40),
        ('PREÇO ORIGINAL', 14),
        ('PREÇO REAJ. 1', 14),
        ('PREÇO REAJ. 2', 14),
        ('PREÇO REAJ. 3', 14),
        ('PREÇO VIGENTE', 14),
    ]

    ihr = ir_start + 1
    for ci, (txt, w) in enumerate(item_hdrs, 1):
        c = ws.cell(row=ihr, column=ci, value=txt)
        c.font = _hdr_font
        c.alignment = _ac
        c.fill = _fill_reajuste
        c.border = _thin
        ws.column_dimensions[get_column_letter(ci)].width = max(
            w, ws.column_dimensions[get_column_letter(ci)].width or 0
        )

    for ri, item in enumerate(itens_vencidos, ihr + 1):
        r = str(ri)
        ws.cell(row=ri, column=1, value=item.get('numero_item', ri - ihr))
        ws.cell(row=ri, column=2, value=(item.get('descricao', '') or '')[:60])

        preco_final = item.get('preco_final')
        if preco_final:
            ws.cell(row=ri, column=3, value=float(preco_final)).number_format = '#,##0.00'

        # G = preço vigente (último preenchido)
        ws.cell(
            row=ri, column=7,
            value=f'=IF(F{r}<>"",F{r},IF(E{r}<>"",E{r},IF(D{r}<>"",D{r},C{r})))'
        ).number_format = '#,##0.00'

        for col in range(1, len(item_hdrs) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = _thin
            cell.font = _dat_font
            if col >= 3:
                cell.number_format = '#,##0.00'
            if col == 7:
                cell.protection = _LOCKED
                cell.fill = _fill_formula
            else:
                cell.protection = _UNLOCKED

    ws.protection.sheet = True
    ws.protection.password = SENHA_PROTECAO
    ws.protection.enable()
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    return ws


# ============================================================
# FUNÇÃO PRINCIPAL: GERAR PLANILHA REAJUSTADA
# ============================================================
def gerar_planilha_reajustada(edital_id, app=None):
    """
    Gera planilha reajustada com itens VENCIDOS de um edital.

    Returns:
        tuple (bytes, stats) ou (None, erro_msg)
    """
    if app is None:
        from flask import current_app
        app = current_app._get_current_object()

    with app.app_context():
        from ..models.database import db, Edital, ItemEditalExtraido

        edital = Edital.query.get(edital_id)
        if not edital:
            logger.error("Edital %d não encontrado para reajustada", edital_id)
            return None, 'Edital não encontrado'

        edital_dict = edital.to_dict()

        # Buscar apenas itens VENCIDOS
        itens_db = ItemEditalExtraido.query.filter_by(
            edital_id=edital_id,
            status_disputa='VENCIDO',
        ).order_by(ItemEditalExtraido.numero_item.asc()).all()

        if not itens_db:
            logger.warning("Nenhum item VENCIDO para edital %d", edital_id)
            return None, 'Nenhum item com status VENCIDO encontrado'

        itens_vencidos = []
        for item in itens_db:
            itens_vencidos.append({
                'numero_item': item.numero_item,
                'descricao': item.descricao,
                'codigo_referencia': item.codigo_referencia,
                'quantidade': float(item.quantidade) if item.quantidade else None,
                'unidade_compra': item.unidade_compra or 'UN',
                'preco_unitario_maximo': float(item.preco_unitario_maximo) if item.preco_unitario_maximo else None,
                'preco_final': float(item.preco_final) if item.preco_final else None,
                'preco_total_final': float(item.preco_total_final) if item.preco_total_final else None,
                'obs_disputa': item.obs_disputa,
                'grupo_lote': item.grupo_lote,
            })

        logger.info(
            "Gerando planilha reajustada: edital=%d, %d itens vencidos",
            edital_id, len(itens_vencidos),
        )

        wb = Workbook()
        _criar_aba_contrato(wb, edital_dict, itens_vencidos)
        _criar_aba_itens_vencidos(wb, itens_vencidos)
        _criar_aba_entregas(wb, itens_vencidos)
        _criar_aba_reajustes(wb, itens_vencidos)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_bytes = buffer.getvalue()

        stats = {
            'itens_vencidos': len(itens_vencidos),
            'valor_total': sum(float(it.get('preco_total_final') or 0) for it in itens_vencidos),
            'tamanho_bytes': len(xlsx_bytes),
        }

        logger.info(
            "Planilha reajustada gerada: edital=%d, %d itens, %d bytes",
            edital_id, len(itens_vencidos), len(xlsx_bytes),
        )

        return xlsx_bytes, stats


def gerar_e_enviar_reajustada(edital_id, app=None):
    """Gera planilha reajustada e envia para o Dropbox."""
    if app is None:
        from flask import current_app
        app = current_app._get_current_object()

    with app.app_context():
        from ..models.database import db, Edital, EditalArquivo
        from . import dropbox_service

        edital = Edital.query.get(edital_id)
        if not edital:
            return {'erro': 'Edital não encontrado'}

        resultado = gerar_planilha_reajustada(edital_id, app)
        xlsx_bytes, stats = resultado

        if not xlsx_bytes:
            return {'erro': stats}  # stats contém a mensagem de erro

        orgao_curto = (edital.orgao_razao_social or 'Edital')[:40].replace('/', '-')
        nome_arquivo = f"REAJUSTADA_{edital.id}_{orgao_curto}.xlsx"

        try:
            pasta = dropbox_service.gerar_pasta_edital(edital)
            dropbox_service.criar_pasta(pasta)
            resultado_upload = dropbox_service.upload_arquivo(
                xlsx_bytes,
                f"{pasta}/{nome_arquivo}",
                nome_arquivo=nome_arquivo,
            )

            if resultado_upload:
                existente = EditalArquivo.query.filter_by(
                    edital_id=edital.id,
                    tipo='reajustada',
                ).first()

                if existente:
                    existente.url_cloudinary = resultado_upload.get('shared_link') or resultado_upload['dropbox_path']
                    existente.nome_arquivo = nome_arquivo
                    existente.tamanho_bytes = resultado_upload['tamanho']
                else:
                    arquivo = EditalArquivo(
                        edital_id=edital.id,
                        tipo='reajustada',
                        nome_arquivo=nome_arquivo,
                        url_cloudinary=resultado_upload.get('shared_link') or resultado_upload['dropbox_path'],
                        tamanho_bytes=resultado_upload['tamanho'],
                        mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )
                    db.session.add(arquivo)

                db.session.commit()
                logger.info("Planilha reajustada enviada: edital=%d, %s", edital_id, nome_arquivo)
                return {'sucesso': True, 'arquivo': nome_arquivo, **stats}
            else:
                return {'erro': 'Falha no upload para Dropbox'}

        except Exception as e:
            db.session.rollback()
            logger.error("Erro enviar reajustada Dropbox edital %d: %s", edital_id, e)
            return {'erro': str(e)}


def disparar_geracao_reajustada_async(edital_id, app):
    """Dispara geração da planilha reajustada em background."""
    thread = Thread(
        target=gerar_e_enviar_reajustada,
        args=(edital_id, app),
        daemon=True,
    )
    thread.start()
    logger.info("Geração de reajustada assíncrona disparada para edital %d", edital_id)
    return thread
