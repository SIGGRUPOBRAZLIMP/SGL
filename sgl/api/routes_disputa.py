"""
SGL - Rotas de Disputa e Planilha Reajustada

NOVOS ENDPOINTS para adicionar ao routes.py (api_bp):
  POST /editais/<id>/registrar-disputa       — lançamento manual dos resultados
  POST /editais/<id>/upload-resultado-disputa — upload da planilha preenchida
  POST /editais/<id>/gerar-reajustada        — gera planilha reajustada (itens VENCIDOS)
  GET  /editais/<id>/resultado-disputa       — consulta resultados registrados

Cole este código no final do routes.py, ANTES da última linha.
"""


# ============================================================
# DISPUTA — REGISTRAR RESULTADOS
# ============================================================

@api_bp.route('/editais/<int:edital_id>/registrar-disputa', methods=['POST'])
@jwt_required()
def registrar_resultado_disputa(edital_id):
    """
    Registra resultado da disputa para itens de um edital.
    
    Body JSON:
    {
        "itens": [
            {
                "item_id": 123,                    // ID do ItemEditalExtraido
                "status_disputa": "VENCIDO",       // VENCIDO, NAO_VENCIDO, DESERTO, FRACASSADO
                "preco_final": 12.50,              // preço unitário efetivo
                "obs_disputa": "negociado"         // opcional
            },
            ...
        ]
    }
    """
    from ..models.database import db, Edital, ItemEditalExtraido

    edital = Edital.query.get(edital_id)
    if not edital:
        return jsonify({'error': 'Edital não encontrado'}), 404

    data = request.get_json()
    itens_input = data.get('itens', [])

    if not itens_input:
        return jsonify({'error': 'Nenhum item enviado'}), 400

    STATUS_VALIDOS = ('VENCIDO', 'NAO_VENCIDO', 'DESERTO', 'FRACASSADO')
    stats = {'atualizados': 0, 'erros': 0, 'vencidos': 0}

    for item_data in itens_input:
        try:
            item_id = item_data.get('item_id')
            status = item_data.get('status_disputa', '').upper().replace(' ', '_').replace('Ã', 'A')

            if status not in STATUS_VALIDOS:
                stats['erros'] += 1
                continue

            item = ItemEditalExtraido.query.filter_by(
                id=item_id, edital_id=edital_id
            ).first()

            if not item:
                stats['erros'] += 1
                continue

            item.status_disputa = status
            item.data_disputa = datetime.now(timezone.utc)
            item.obs_disputa = item_data.get('obs_disputa')

            if status == 'VENCIDO':
                preco = item_data.get('preco_final')
                if preco is not None:
                    item.preco_final = float(preco)
                    qtd = float(item.quantidade) if item.quantidade else 0
                    item.preco_total_final = float(preco) * qtd
                stats['vencidos'] += 1

            stats['atualizados'] += 1

        except Exception as e:
            current_app.logger.error('Erro registrar disputa item %s: %s', item_data, e)
            stats['erros'] += 1

    db.session.commit()

    # Atualizar status do edital
    total_vencidos = ItemEditalExtraido.query.filter_by(
        edital_id=edital_id, status_disputa='VENCIDO'
    ).count()
    total_disputados = ItemEditalExtraido.query.filter(
        ItemEditalExtraido.edital_id == edital_id,
        ItemEditalExtraido.status_disputa.isnot(None),
    ).count()

    if total_vencidos > 0:
        edital.status = 'ganho_parcial' if total_vencidos < total_disputados else 'ganho_total'
    elif total_disputados > 0:
        edital.status = 'perdido'

    db.session.commit()

    return jsonify({
        'sucesso': True,
        **stats,
        'total_vencidos': total_vencidos,
        'status_edital': edital.status,
    })


# ============================================================
# DISPUTA — UPLOAD PLANILHA PREENCHIDA
# ============================================================

@api_bp.route('/editais/<int:edital_id>/upload-resultado-disputa', methods=['POST'])
@jwt_required()
def upload_resultado_disputa(edital_id):
    """
    Recebe upload da planilha de cotação preenchida com resultados da disputa.
    Lê as colunas STATUS DISPUTA e PREÇO FINAL e salva no banco.
    
    Espera multipart/form-data com campo 'arquivo' (XLSX).
    """
    from ..models.database import db, Edital, ItemEditalExtraido

    edital = Edital.query.get(edital_id)
    if not edital:
        return jsonify({'error': 'Edital não encontrado'}), 404

    if 'arquivo' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado. Use campo "arquivo".'}), 400

    arquivo = request.files['arquivo']
    if not arquivo.filename.endswith('.xlsx'):
        return jsonify({'error': 'Formato inválido. Envie um arquivo .xlsx'}), 400

    try:
        import openpyxl

        wb = openpyxl.load_workbook(arquivo, data_only=True)

        # Encontrar a aba "Cotação"
        if 'Cotação' in wb.sheetnames:
            ws = wb['Cotação']
        elif 'Cotacao' in wb.sheetnames:
            ws = wb['Cotacao']
        else:
            # Tentar segunda aba
            ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

        # Detectar colunas pelo cabeçalho (row 1)
        col_item = None
        col_status_disputa = None
        col_preco_final = None
        col_obs_disputa = None

        for col in range(1, ws.max_column + 1):
            hdr = str(ws.cell(row=1, column=col).value or '').upper().replace('\n', ' ')
            if 'ITEM' == hdr.strip():
                col_item = col
            elif 'STATUS' in hdr and 'DISPUTA' in hdr:
                col_status_disputa = col
            elif 'PREÇO' in hdr and 'FINAL' in hdr and 'TOTAL' not in hdr:
                col_preco_final = col
            elif 'OBS' in hdr and 'DISPUTA' in hdr:
                col_obs_disputa = col

        if not col_status_disputa:
            return jsonify({
                'error': 'Coluna "STATUS DISPUTA" não encontrada na planilha. '
                         'Verifique se é a planilha de cotação correta.'
            }), 400

        # Buscar itens do edital indexados por numero_item
        itens_db = {
            i.numero_item: i
            for i in ItemEditalExtraido.query.filter_by(edital_id=edital_id).all()
            if i.numero_item
        }

        STATUS_MAP = {
            'VENCIDO': 'VENCIDO',
            'NÃO VENCIDO': 'NAO_VENCIDO',
            'NAO VENCIDO': 'NAO_VENCIDO',
            'NÃO_VENCIDO': 'NAO_VENCIDO',
            'NAO_VENCIDO': 'NAO_VENCIDO',
            'DESERTO': 'DESERTO',
            'FRACASSADO': 'FRACASSADO',
        }

        stats = {'lidos': 0, 'atualizados': 0, 'vencidos': 0, 'sem_match': 0}

        for row in range(2, ws.max_row + 1):
            status_val = ws.cell(row=row, column=col_status_disputa).value
            if not status_val:
                continue

            status_str = str(status_val).strip().upper()
            status_norm = STATUS_MAP.get(status_str)
            if not status_norm:
                continue

            stats['lidos'] += 1

            # Encontrar o item correspondente
            num_item = ws.cell(row=row, column=col_item).value if col_item else None
            item_db = None

            if num_item is not None:
                try:
                    item_db = itens_db.get(int(num_item))
                except (ValueError, TypeError):
                    pass

            # Fallback: buscar pela posição (row - 2 = index)
            if not item_db:
                idx = row - 2
                for it in itens_db.values():
                    if it.numero_item == idx + 1:
                        item_db = it
                        break

            if not item_db:
                stats['sem_match'] += 1
                continue

            item_db.status_disputa = status_norm
            item_db.data_disputa = datetime.now(timezone.utc)

            if col_preco_final:
                preco = ws.cell(row=row, column=col_preco_final).value
                if preco is not None:
                    try:
                        item_db.preco_final = float(preco)
                        qtd = float(item_db.quantidade) if item_db.quantidade else 0
                        item_db.preco_total_final = float(preco) * qtd
                    except (ValueError, TypeError):
                        pass

            if col_obs_disputa:
                obs = ws.cell(row=row, column=col_obs_disputa).value
                if obs:
                    item_db.obs_disputa = str(obs)

            stats['atualizados'] += 1
            if status_norm == 'VENCIDO':
                stats['vencidos'] += 1

        db.session.commit()

        # Atualizar status do edital
        total_vencidos = ItemEditalExtraido.query.filter_by(
            edital_id=edital_id, status_disputa='VENCIDO'
        ).count()

        if total_vencidos > 0:
            edital.status = 'ganho_parcial'
        db.session.commit()

        return jsonify({
            'sucesso': True,
            **stats,
            'total_vencidos_banco': total_vencidos,
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Erro upload resultado disputa edital %d: %s', edital_id, e)
        return jsonify({'error': f'Erro ao processar planilha: {str(e)}'}), 500


# ============================================================
# DISPUTA — CONSULTAR RESULTADOS
# ============================================================

@api_bp.route('/editais/<int:edital_id>/resultado-disputa', methods=['GET'])
@jwt_required()
def consultar_resultado_disputa(edital_id):
    """Retorna itens do edital com resultado da disputa."""
    from ..models.database import ItemEditalExtraido

    itens = ItemEditalExtraido.query.filter_by(edital_id=edital_id).order_by(
        ItemEditalExtraido.numero_item.asc()
    ).all()

    resultado = {
        'edital_id': edital_id,
        'total_itens': len(itens),
        'vencidos': 0,
        'nao_vencidos': 0,
        'desertos': 0,
        'fracassados': 0,
        'sem_resultado': 0,
        'valor_total_vencido': 0,
        'itens': [],
    }

    for item in itens:
        d = item.to_dict()
        resultado['itens'].append(d)

        sd = item.status_disputa
        if sd == 'VENCIDO':
            resultado['vencidos'] += 1
            resultado['valor_total_vencido'] += float(item.preco_total_final or 0)
        elif sd == 'NAO_VENCIDO':
            resultado['nao_vencidos'] += 1
        elif sd == 'DESERTO':
            resultado['desertos'] += 1
        elif sd == 'FRACASSADO':
            resultado['fracassados'] += 1
        else:
            resultado['sem_resultado'] += 1

    return jsonify(resultado)


# ============================================================
# PLANILHA REAJUSTADA — GERAR
# ============================================================

@api_bp.route('/editais/<int:edital_id>/gerar-reajustada', methods=['POST'])
@jwt_required()
def gerar_reajustada_edital(edital_id):
    """
    Gera planilha reajustada com itens VENCIDOS e envia ao Dropbox.
    Pré-requisito: ter registrado resultado da disputa com pelo menos 1 item VENCIDO.
    """
    from ..models.database import Edital, ItemEditalExtraido

    edital = Edital.query.get(edital_id)
    if not edital:
        return jsonify({'error': 'Edital não encontrado'}), 404

    # Verificar se tem itens vencidos
    vencidos = ItemEditalExtraido.query.filter_by(
        edital_id=edital_id, status_disputa='VENCIDO'
    ).count()

    if vencidos == 0:
        return jsonify({
            'error': 'Nenhum item VENCIDO encontrado. Registre o resultado da disputa primeiro.'
        }), 400

    try:
        from ..services.planilha_reajustada_service import gerar_e_enviar_reajustada
        resultado = gerar_e_enviar_reajustada(edital_id, current_app._get_current_object())

        if resultado.get('erro'):
            return jsonify(resultado), 400

        return jsonify(resultado)

    except Exception as e:
        current_app.logger.error('Erro gerar reajustada edital %d: %s', edital_id, e)
        return jsonify({'error': str(e)}), 500
